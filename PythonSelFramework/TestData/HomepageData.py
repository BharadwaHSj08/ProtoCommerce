import openpyxl


class HomepageData:

    test_HomepageData = [{"firstname": "sai", "email": "sai@gmail.com", "Gender": "Male"},{"firstname":"Aparna","email":"Aparna@gmail.com","Gender":"Female"}]

    #In order to call the methods in a call without creating an object for that class,we can use staticmethod
    #we can directly call the method with the class -->HomepageData.getTestData
    @staticmethod
    def getTestData(test_case_name):
        # Location of file
        book = openpyxl.load_workbook("C:\\Users\\Personal\\OneDrive\\Desktop\\PythonDataDriven.xlsx")

        # To open the sheet which we need to work on
        sheet = book.active
        Dict = {}
        for i in range(2, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i,column=1).value == test_case_name:  # if we only want TC3 then it will not go inside loop and scans columnsfor that
                for j in range(2, sheet.max_column + 1):  # to get columns
                    # print(sheet.cell(row=i,column=j).value)
                    # Dictionary["firstname"]="Hema" in this way
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]