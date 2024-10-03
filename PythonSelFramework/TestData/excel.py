import openpyxl

#Location of file
book =openpyxl.load_workbook("C:\\Users\\Personal\\OneDrive\\Desktop\\PythonDataDriven.xlsx")

#To open the sheet which we need to work on
sheet = book.active

#to get the value of row and colmns
cell = sheet.cell(row=2,column=2)
print(cell.value)

#writing the value
sheet.cell(row=2,column=2).value = "Bharat"
print(sheet.cell(row=2,column=2).value)

#find max rows
print(sheet.max_row)
print(sheet.max_column)

Dict = {}
for i in range(2,sheet.max_row+1): #to get rows
    if  sheet.cell(row=i,column=1).value == "TC2": #if we only want TC3 then it will not go inside loop and scans columnsfor that
        for j in range(2,sheet.max_column+1): #to get columns
            #print(sheet.cell(row=i,column=j).value)
            #Dictionary["firstname"]="Hema" in this way
            Dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value

print(Dict)